"""Regenera el xlsx completo desde SQLite.

Una hoja por nicho, mas "Todos" para filtrar y "Descartados" para auditar lo que
el clasificador saco. El nicho de cada hoja lo decide la categoria real de Maps
(ver clasificador.py), no la busqueda que encontro al negocio.
"""
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.clasificador import clasificar
from modules.calificar import calificar
from config.generar_busquedas import PRIORIDAD

# Orden pensado para llamar: primero a quien llamas y a que numero, despues
# donde esta y que tan grande es.
COLUMNAS = ['prioridad', 'porque', 'nombre', 'telefono', 'categoria', 'ciudad',
            'web', 'direccion', 'rating', 'resenas', 'url']
TITULOS = {'rubro': 'Rubro', 'nicho': 'Nicho', 'nombre': 'Negocio',
           'telefono': 'Teléfono', 'categoria': 'Categoría', 'ciudad': 'Ciudad',
           'web': 'Sitio web', 'direccion': 'Dirección', 'rating': 'Puntaje',
           'resenas': 'Reseñas', 'url': 'Link en Maps', 'motivo': 'Motivo del descarte',
           'prioridad': 'Prioridad', 'porque': 'Por qué'}
ANCHOS = {'rubro': 20, 'nicho': 22, 'nombre': 42, 'telefono': 16, 'categoria': 26,
          'ciudad': 18, 'web': 32, 'direccion': 44, 'rating': 8, 'resenas': 9,
          'url': 14, 'motivo': 40, 'prioridad': 10, 'porque': 34}


def _nombre_hoja(texto, usados):
    hoja = re.sub(r'[\[\]:*?/\\]', '-', str(texto)).strip()[:31] or "Sin nicho"
    base, n = hoja, 2
    while hoja in usados:
        hoja = f"{base[:28]}_{n}"
        n += 1
    usados.add(hoja)
    return hoja


def _rubro_de(nicho, mapa):
    return mapa.get(nicho, 'Otros servicios')


# Algunos negocios publican el numero sin codigo de area ("4450405"), que no se
# puede discar desde otra ciudad. Se completa con el de la ciudad donde esta.
AREA = {'Rosario': '0341', 'Buenos Aires': '011', 'Córdoba': '0351',
        'Mendoza': '0261', 'La Plata': '0221', 'Mar del Plata': '0223',
        'San Miguel de Tucumán': '0381', 'Salta': '0387', 'Santa Fe': '0342',
        'Neuquén': '0299', 'Bahía Blanca': '0291', 'Paraná': '0343'}


def _completar_tel(telefono, ciudad):
    # Con gl=AR un negocio argentino da formato local (0341...). Un prefijo
    # internacional significa que Maps devolvio otro pais: hubo 91 de Córdoba,
    # España. No sirven como lead argentino.
    if str(telefono or '').startswith('+'):
        return ''
    t = re.sub(r'\D', '', str(telefono or ''))
    if not t or t.startswith('0'):
        return telefono
    if len(set(t)) <= 2 or len(t) < 6:      # 11111111, 0000, basura
        return ''
    area = AREA.get(ciudad)
    return f"{area}{t}" if area and len(t) <= 8 else telefono


def exportar(con, archivo_salida="output/Organizate.xlsx",
             ruta_busquedas="config/busquedas.xlsx"):
    """Solo exporta negocios CON telefono. Los que no tienen quedan en la base:
    si en una pasada posterior aparece el numero, entran solos."""
    df = pd.read_sql_query(
        "SELECT nombre, telefono, categoria, ciudad, direccion, web, rating, "
        "resenas, url, nicho FROM negocios WHERE telefono != ''", con)
    if df.empty:
        print("No hay negocios con telefono para exportar.")
        return

    try:
        b = pd.read_excel(ruta_busquedas)[['Nicho', 'Rubro']].drop_duplicates()
        mapa_rubro = dict(zip(b['Nicho'], b['Rubro']))
    except Exception:
        mapa_rubro = {}

    decidido = df.apply(
        lambda f: clasificar(f['categoria'], f['nicho'], f['nombre']), axis=1)
    df['nicho'] = [d[0] for d in decidido]
    df['motivo'] = [d[1] for d in decidido]

    df['telefono'] = [_completar_tel(t, c) for t, c in zip(df['telefono'], df['ciudad'])]
    df = df[df['telefono'] != '']

    descartados = df[df['nicho'].isna()].copy()
    df = df[df['nicho'].notna()].copy()

    # Un mismo telefono = una sola llamada. Maps a veces tiene dos fichas del
    # mismo local (distinto place_id, asi que la dedup de la base no las junta).
    # Se queda la de mas reseñas, que suele ser la ficha viva.
    df['_peso'] = pd.to_numeric(df['resenas'], errors='coerce').fillna(0)
    antes = len(df)
    df = (df.sort_values('_peso', ascending=False)
            .drop_duplicates(subset=['telefono'], keep='first')
            .drop(columns='_peso'))
    repetidos = antes - len(df)
    df['rubro'] = df['nicho'].map(lambda n: _rubro_de(n, mapa_rubro))

    # A quien llamar primero. La prioridad del nicho sale de generar_busquedas,
    # que es donde ya decidiste que un gimnasio no vende turnos: sin esto los
    # 8323 negocios de nichos apagados que se scrapearon antes de apagarlos
    # siguen apareciendo en la planilla y los vendedores los llaman igual.
    calif = df.apply(lambda f: calificar(PRIORIDAD.get(f['nicho'], 3), f['telefono'],
                                         f['web'], f['resenas'], f['rating']), axis=1)
    df['prioridad'] = [c[0] for c in calif]
    df['porque'] = [c[1] for c in calif]

    df = df.sort_values(['rubro', 'nicho', 'prioridad', 'ciudad', 'nombre'])

    os.makedirs(os.path.dirname(archivo_salida), exist_ok=True)
    usados, hojas = set(), []
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        # "Todos" primero: es la hoja para filtrar y cruzar.
        cols_todos = ['rubro', 'nicho'] + COLUMNAS
        hoja = _nombre_hoja("Todos", usados)
        df[cols_todos].rename(columns=TITULOS).to_excel(writer, sheet_name=hoja, index=False)
        hojas.append((hoja, cols_todos, len(df)))

        for nicho, grupo in df.groupby('nicho', sort=True):
            hoja = _nombre_hoja(nicho, usados)
            grupo[COLUMNAS].rename(columns=TITULOS).to_excel(
                writer, sheet_name=hoja, index=False)
            hojas.append((hoja, COLUMNAS, len(grupo)))

        if not descartados.empty:
            cols_desc = ['nombre', 'telefono', 'categoria', 'ciudad', 'motivo']
            hoja = _nombre_hoja("Descartados", usados)
            descartados[cols_desc].rename(columns=TITULOS).to_excel(
                writer, sheet_name=hoja, index=False)
            hojas.append((hoja, cols_desc, len(descartados)))

    _formatear(archivo_salida, hojas)
    sin_tel = con.execute("SELECT COUNT(*) FROM negocios WHERE telefono = ''").fetchone()[0]
    print(f"Excel: {len(df)} negocios en {len(hojas) - 2} nichos -> {archivo_salida}")
    print(f"       {len(descartados)} descartados por no vender turnos (hoja Descartados)")
    print(f"       {repetidos} unificados por compartir telefono")
    print(f"       {sin_tel} sin telefono quedaron en la base, fuera del Excel")


def _formatear(archivo, hojas):
    try:
        wb = load_workbook(archivo)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", start_color="2F5597")
        fino = Side(style='thin', color="D0D0D0")
        borde = Border(left=fino, right=fino, top=fino, bottom=fino)

        for hoja, claves, _ in hojas:
            ws = wb[hoja]
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions

            for i, clave in enumerate(claves, start=1):
                ws.column_dimensions[get_column_letter(i)].width = ANCHOS.get(clave, 20)
                if clave == 'telefono':
                    # Formato texto: sin esto Excel lee 03416359476 como numero y
                    # se come el 0 inicial, que hace inservible el numero.
                    for fila in range(2, ws.max_row + 1):
                        ws.cell(row=fila, column=i).number_format = '@'
                elif clave == 'url':
                    # La url queda como valor, no como texto "ver en Maps": es lo
                    # que despues se copia a Google Sheets, donde el texto solo no
                    # sirve para nada. El hipervinculo la deja clickeable igual.
                    for fila in range(2, ws.max_row + 1):
                        celda = ws.cell(row=fila, column=i)
                        if celda.value:
                            celda.hyperlink = celda.value
                            celda.font = Font(color="0563C1", underline="single")

            for celda in ws[1]:
                celda.font, celda.fill, celda.border = header_font, header_fill, borde
                celda.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22
        wb.save(archivo)
    except Exception as e:
        print(f"Aviso: no se pudo formatear el Excel: {e}")
