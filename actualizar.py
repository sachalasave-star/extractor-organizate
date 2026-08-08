"""Trae los negocios nuevos al Excel donde trabajan los vendedores, SIN tocar
lo que ya anotaron.

    python actualizar.py

output/Organizate.xlsx lo regenera el scraper entero en cada corrida: escribir
ahi es perder el trabajo. Los vendedores usan output/Seguimiento.xlsx, que este
script solo AGREGA filas al final. Las que ya estaban no se tocan ni se
reordenan, aunque el negocio haya cambiado de nombre o de nicho en Maps.

Se adapta a las columnas que tenga el Seguimiento: si le agregaste "Vendedor",
"Estado" o lo que sea, se respetan y quedan vacias en las filas nuevas.
"""
import os
import shutil
import sys
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

GENERADO = "output/Organizate.xlsx"
SEGUIMIENTO = "output/Seguimiento.xlsx"
CLAVE = "Teléfono"          # identifica al negocio: unico tras la dedup del export

# Columnas de gestion que agrega la primera vez. Podes renombrarlas, agregar mas
# o borrarlas desde Excel: el script respeta lo que encuentre en el archivo.
GESTION = ["Estado", "Vendedor", "Fecha de llamada", "Notas"]
ESTADOS = ['Sin llamar', 'Llamado', 'No contesta', 'Volver a llamar',
           'No llamar mas', 'Interesado', 'Vendido']


def _hoja_datos(archivo):
    """La hoja 'Todos' del generado trae todo junto; si no esta, la primera."""
    x = pd.ExcelFile(archivo)
    return "Todos" if "Todos" in x.sheet_names else x.sheet_names[0]


def actualizar(generado=GENERADO, seguimiento=SEGUIMIENTO):
    if not os.path.exists(generado):
        sys.exit(f"No existe {generado}. Corre el scraper primero.")

    nuevos = pd.read_excel(generado, sheet_name=_hoja_datos(generado), dtype={CLAVE: str})
    nuevos = nuevos[nuevos[CLAVE].notna()]

    if not os.path.exists(seguimiento):
        salida = nuevos.copy()
        for col in GESTION:
            salida[col] = ""
        salida.loc[:, "Estado"] = ESTADOS[0]
        _escribir(salida, seguimiento, es_nuevo=True)
        print(f"Creado {seguimiento} con {len(salida)} negocios.")
        print(f"Columnas para trabajar: {', '.join(GESTION)}")
        return

    # dtype=str en todo: una columna de gestion vacia se lee como numerica y
    # despues rechaza el texto que escriba el vendedor.
    viejos = pd.read_excel(seguimiento, dtype=str)
    if CLAVE not in viejos.columns:
        sys.exit(f"{seguimiento} no tiene columna '{CLAVE}'. No puedo saber que "
                 f"negocio es cual sin ella.")

    ya_estan = set(viejos[CLAVE].dropna())
    faltan = nuevos[~nuevos[CLAVE].isin(ya_estan)].copy()
    if faltan.empty:
        print(f"Sin novedades: los {len(ya_estan)} negocios ya estaban en {seguimiento}.")
        return

    # Copia de seguridad antes de escribir: el archivo tiene trabajo humano adentro.
    respaldo = seguimiento.replace('.xlsx', f"_backup_{datetime.now():%Y%m%d_%H%M}.xlsx")
    shutil.copy2(seguimiento, respaldo)

    # Las columnas mandan las del seguimiento: si el usuario agrego o renombro
    # alguna, se conserva su planilla tal cual y las nuevas filas se acomodan.
    for col in viejos.columns:
        if col not in faltan.columns:
            faltan[col] = ESTADOS[0] if col == "Estado" else ""
    faltan = faltan[viejos.columns]

    _escribir(pd.concat([viejos, faltan], ignore_index=True), seguimiento)
    print(f"Agregados {len(faltan)} negocios nuevos a {seguimiento} "
          f"({len(ya_estan)} ya estaban, intactos).")
    print(f"Respaldo del anterior: {respaldo}")


def _escribir(df, destino, es_nuevo=False):
    df.to_excel(destino, index=False, sheet_name="Seguimiento")
    try:
        wb = load_workbook(destino)
        ws = wb["Seguimiento"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for celda in ws[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", start_color="2F5597")
            celda.alignment = Alignment(horizontal='center', vertical='center')
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            ancho = 40 if col in ("Negocio", "Dirección", "Notas") else 18
            ws.column_dimensions[letra].width = ancho
            if col == CLAVE:
                for fila in range(2, ws.max_row + 1):
                    ws.cell(row=fila, column=i).number_format = '@'
        wb.save(destino)
    except Exception as e:
        print(f"Aviso: no se pudo formatear ({e}). Los datos igual se guardaron.")


def demo():
    """Lo unico que importa: agregar negocios nuevos NO puede pisar anotaciones."""
    import tempfile
    d = tempfile.mkdtemp()
    gen, seg = os.path.join(d, 'g.xlsx'), os.path.join(d, 's.xlsx')

    pd.DataFrame({'Negocio': ['A', 'B'], CLAVE: ['0341111', '0341222']}
                 ).to_excel(gen, index=False, sheet_name='Todos')
    actualizar(gen, seg)

    # El vendedor trabaja: anota sobre el negocio A
    t = pd.read_excel(seg, dtype=str)
    t.loc[t[CLAVE] == '0341111', ['Estado', 'Vendedor', 'Notas']] = \
        ['Vendido', 'Sacha', 'cerro 3 sucursales']
    t.to_excel(seg, index=False, sheet_name='Seguimiento')

    # El scraper encuentra un negocio mas
    pd.DataFrame({'Negocio': ['A', 'B', 'C'], CLAVE: ['0341111', '0341222', '0341333']}
                 ).to_excel(gen, index=False, sheet_name='Todos')
    actualizar(gen, seg)

    r = pd.read_excel(seg, dtype={CLAVE: str})
    fila = r[r[CLAVE] == '0341111'].iloc[0]
    assert fila['Estado'] == 'Vendido', f"se piso el estado: {fila['Estado']!r}"
    assert fila['Vendedor'] == 'Sacha', 'se piso el vendedor'
    assert fila['Notas'] == 'cerro 3 sucursales', 'se pisaron las notas'
    assert len(r) == 3, f'esperaba 3 filas, hay {len(r)}'
    assert (r[CLAVE] == '0341333').any(), 'no se agrego el negocio nuevo'
    print('OK  actualizar: las anotaciones sobreviven y entran los nuevos')


if __name__ == "__main__":
    if '--test' in sys.argv:
        demo()
    else:
        actualizar()
